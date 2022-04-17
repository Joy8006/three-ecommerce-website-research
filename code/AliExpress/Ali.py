import openpyxl
from selenium import webdriver
import time
import pandas as pd
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('disable-notifications')

driver = webdriver.Chrome('chromedriver.exe', options=chrome_options)

wb_obj = openpyxl.load_workbook("AliExpressStationaryLinks.xlsx")
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row

driver.maximize_window()

Title = []
Price = []
Rating = []
Review = []
Order = []
Link = []

for i in range(1, m_row):
    cell_obj = sheet_obj.cell(row=i+1, column=2)
    # print(cell_obj.value)
    driver.get(cell_obj.value)
    # time.sleep(1)
    try:
        
        title = driver.find_element_by_class_name("product-title-text").text
        
    except NoSuchElementException:
        title = "N/A"
    try:

        price = driver.find_element_by_class_name("product-price-value").text.replace("BDT", "").strip()
        
    except NoSuchElementException:
        price = "0"
        

    try:
        rating = driver.find_element_by_class_name("overview-rating-average").text
        
    except NoSuchElementException:
        rating = "0"
        

    try:
        review = driver.find_element_by_class_name("product-reviewer-reviews").text.replace("Reviews", "").replace("Review", "")
        
    except NoSuchElementException:
        review = "0"
        
    try:
        order = driver.find_element_by_class_name("product-reviewer-sold").text.replace("orders", "").replace("order", "")
    except NoSuchElementException:
        order = "0"
    Title.append(title)
    Price.append(price)
    Rating.append(rating)
    Review.append(review)
    Order.append(order)
    Link.append(cell_obj.value)
    print("{} / {}".format(i, m_row-1))

list_of_tuples = list(zip(Title, Price, Rating, Review, Order, Link))
df = pd.DataFrame(list_of_tuples, columns= ['Product Title', 'Product Price','Rating', 'Reviews', 'Orders', 'Product Link'])
df.to_excel("AliExpressOverview.xlsx")
driver.close()
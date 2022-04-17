# from trendhol import title, company, price
from selenium import webdriver
import time
import pandas as pd
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
import openpyxl
from googletrans import Translator

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('disable-notifications')

driver = webdriver.Chrome('chromedriver.exe', options=chrome_options)


wb_obj = openpyxl.load_workbook("trendyol.xlsx")
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
translator = Translator()
ali = []
titleAli = []
titleEng = []

for i in range(1, m_row):
    cell_obj = sheet_obj.cell(row = i+1, column = 3)
    print(translator.translate(cell_obj.value).text)
    search = translator.translate(cell_obj.value, src='tr', dest='en').text
    if(i==1):

        driver.get("https://www.aliexpress.com/?spm=a2g0o.productlist.1000002.1.2d0462bdpOVKN3")
        driver.maximize_window()
        driver.find_element_by_xpath("/html/body/div[3]/div[4]/div/div[3]/form/div[2]/input").send_keys(search)
        driver.find_element_by_xpath("/html/body/div[3]/div[4]/div/div[3]/form/div[2]/input").send_keys(Keys.ENTER)
        time.sleep(2)
        xx = driver.find_element_by_xpath("/html/body/div[3]/div[1]/div/div[2]/div[2]/div/div[2]/a[1]/div[2]/div[2]")
        yy = xx.text
        titleEng.append(search)
        titleAli.append(cell_obj.value)
        ali.append(yy)
        
    elif(i>1):
        driver.find_element_by_xpath("/html/body/div[2]/div[4]/div/div[3]/form/div[2]/input").clear()
        driver.find_element_by_xpath("/html/body/div[2]/div[4]/div/div[3]/form/div[2]/input").send_keys(search)
        driver.find_element_by_xpath("/html/body/div[2]/div[4]/div/div[3]/form/div[2]/input").send_keys(Keys.ENTER)
        try:

            xx = driver.find_element_by_xpath("/html/body/div[3]/div[1]/div/div[2]/div[2]/div/div[2]/a[1]/div[2]/div[2]")
            yy = xx.text
        except NoSuchElementException:
            yy = "Not Found"
        titleAli.append(cell_obj.value)
        ali.append(yy)
        titleEng.append(search)
    

driver.close()


list_of_tuples = list(zip(titleAli,titleEng, ali))

df = pd.DataFrame(list_of_tuples,
                  columns = ['Title Ali','Title English', 'Ali Express Price'])


df.to_excel('AliExpress.xlsx')

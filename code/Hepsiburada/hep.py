from selenium import webdriver
import time
import pandas as pd
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
# from googletrans import Translator

chrome_options = webdriver.ChromeOptions()
chrome_options.add_argument('disable-notifications')

driver = webdriver.Chrome('chromedriver.exe', options=chrome_options)
driver.maximize_window()

wb_obj = openpyxl.load_workbook("hasiburada.xlsx")
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row

Title = []
Brand = []
Price = []
Rating = []
Evaluation = []
Link = []


for i in range(1, m_row):
    cell_obj = sheet_obj.cell(row = i+1, column = 2)
    driver.get(cell_obj.value)

    
    try:

        title= driver.find_element_by_css_selector("#product-name").text
    except NoSuchElementException:
        title = "N/A"
    try:

        brand = driver.find_element_by_css_selector("body > div.wrapper > main > div.product-detail-module > section.detail-main > div.container.contain-lg-4.contain-md-4.contain-sm-1.fluid > div > div.productDetailRight.col.lg-2.md-2.sm-1.grid-demo-fluid > div.product-information.col.lg-5.sm-1 > span").text
    except NoSuchElementException:
        brand = "N/A"
    try:

        price = driver.find_element_by_css_selector("#offering-price").text.replace("TL", "")
    except NoSuchElementException:
        price = "0"

    try:
        rating = driver.find_element_by_class_name("rating-star").text
    except NoSuchElementException:
        rating = "0"

    try:
        evaluation = driver.find_element_by_id("comments-container").text.replace("Değerlendirme", "")
    except NoSuchElementException:
        evaluation = "0"
    Title.append(title)
    Brand.append(brand)
    Price.append(price)
    Rating.append(rating)
    Evaluation.append(evaluation)
    Link.append(cell_obj.value)
    print("{} / {}".format(i,m_row))

list_of_tuples = list(zip(Title, Brand, Price, Rating, Evaluation, Link))
df = pd.DataFrame(list_of_tuples, columns= ['Product Title', 'Brand Name', 'Price', 'Rating', 'Evaluations', 'Product Link'])
df.to_excel("HasiburadaStationary.xlsx")

driver.close()
